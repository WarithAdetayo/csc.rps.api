from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
import openpyxl
from appdata.models import CourseRegistration, Course, UploadHistory, SessionRegistration, AcademicSession
from appdata.models.enums.choices import ParseStatus
from django.contrib.auth.models import AnonymousUser
from drf_spectacular.utils import extend_schema_view, extend_schema
from api.system_administration.serializers import UploadHistorySerializer, UploadResponseSerializer
from django.db import transaction

@extend_schema_view(
    get=extend_schema(
        summary='Generate score sheet template',
        description='Generate an Excel template for uploading scores with matric number, CA score, exam score, and total for a given session and course',
        tags=['Score Sheet Upload'],
        responses={
            200: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            400: {'description': 'Invalid session or course'}
        }
    )
)
class GenerateTemplateView(APIView):
    def get(self, request, session_id, course_id):
        try:
            session = AcademicSession.objects.get(academic_session_id=session_id)
            course = Course.objects.get(course_id=course_id)
        except (AcademicSession.DoesNotExist, Course.DoesNotExist):
            return Response({"error": "Invalid session or course"}, status=status.HTTP_400_BAD_REQUEST)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{course.course_code} Scores"
        ws.append(["matric_number", "course_code", "ca_score", "exam_score", "total"])
        ws.append(["", course.course_code, "", "", ""])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={course.course_code}_template.xlsx"
        wb.save(response)
        return response

@extend_schema_view(
    post=extend_schema(
        summary='Upload score sheet',
        description='Upload an Excel file with student scores (CA and exam) for a session and course. Rejects the sheet if any errors occur.',
        tags=['Score Sheet Upload'],
        request={'multipart/form-data': {'file': 'file'}},
        responses={
            201: UploadResponseSerializer,
            400: UploadResponseSerializer
        }
    )
)
class UploadScoreSheetView(APIView):
    def post(self, request, session_id, course_id):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            session = AcademicSession.objects.get(academic_session_id=session_id)
            course = Course.objects.get(course_id=course_id)
        except (AcademicSession.DoesNotExist, Course.DoesNotExist):
            return Response({"error": "Invalid session or course"}, status=status.HTTP_400_BAD_REQUEST)

        upload = UploadHistory(session=session, course=course, uploaded_file=file)
        if isinstance(request.user, AnonymousUser):
            upload.created_by_user = "system"
            upload.last_modified_by_user = "system"
        else:
            upload.created_by_user = request.user.user_id
            upload.last_modified_by_user = request.user.user_id
        upload.save()

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        errors = []

        # Validate all rows before any updates
        updates = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            matric_number, course_code, ca_score, exam_score, total = row
            if not all([matric_number, course_code, ca_score is not None, exam_score is not None, total is not None]):
                errors.append(f"Missing data for student {matric_number}")
                continue

            # Validate total
            if ca_score + exam_score != total:
                errors.append(f"Total mismatch for {matric_number}: CA ({ca_score}) + Exam ({exam_score}) != {total}")
                continue

            try:
                session_reg = SessionRegistration.objects.get(
                    student__matric_number=matric_number,
                    session=session
                )
                registration = CourseRegistration.objects.get(
                    session_registration=session_reg,
                    course=course
                )
                updates.append((registration, ca_score, exam_score))
            except SessionRegistration.DoesNotExist:
                errors.append(f"Student {matric_number} not registered for session {session}")
            except CourseRegistration.DoesNotExist:
                errors.append(f"Course registration not found for student {matric_number} and course {course.course_code}")
            except Exception as e:
                errors.append(str(e))

        # If any errors, reject the sheet
        if errors:
            upload.parse_status = ParseStatus.FAILED
            upload.save()
            return Response({"message": "Upload failed", "upload_id": upload.upload_history_id, "errors": errors}, status=status.HTTP_400_BAD_REQUEST)

        # Apply updates within a transaction
        with transaction.atomic():
            for registration, ca_score, exam_score in updates:
                registration.ca_score = ca_score
                registration.exam_score = exam_score
                if isinstance(request.user, AnonymousUser):
                    registration.last_modified_by_user = "system"
                else:
                    registration.last_modified_by_user = request.user.user_id
                registration.save()

            upload.parse_status = ParseStatus.SUCCESSFUL
            upload.save()

        return Response({"message": "Upload successful", "upload_id": upload.upload_history_id}, status=status.HTTP_201_CREATED)

@extend_schema_view(
    get=extend_schema(
        summary='List upload history',
        description='Retrieve the history of all score sheet uploads',
        tags=['Score Sheet Upload'],
        responses={200: UploadHistorySerializer(many=True)}
    )
)
class UploadHistoryView(ListAPIView):
    queryset = UploadHistory.objects.all()
    serializer_class = UploadHistorySerializer